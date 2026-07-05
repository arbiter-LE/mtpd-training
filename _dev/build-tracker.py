"""
Arbiter LE — Work Board (project-management tracker)
Regenerates 2026-06-22-arbiter-le-tracker.xlsx in the repo root.

This is the internal work board (NOT the sales lead pipeline — that's
_dev/build-lead-tracker.py). Caught by *.xlsx in .gitignore, so it never
commits and never deploys.

The output file is self-maintaining: status/priority/bucket dropdowns,
conditional formatting, and the Dashboard formulas all live inside the
workbook and keep working as Andrew edits in Excel. Re-running this script
RESEEDS the task rows to the snapshot below — only rerun to re-scaffold,
not as part of the weekly edit loop.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ── Brand tokens (navy / gold / steel) ──
NAVY="0A1828"; GOLD="C8902A"; STEEL="3C6478"
LIGHT="EFF2F4"; WHITE="FFFFFF"
RED_FILL="F8D7DA"; GOLD_FILL="FCEFC8"; GREEN_FILL="D7E8DC"; BLUE_FILL="DCE6EC"
RED_TXT="9C2A2A"; GREEN_TXT="2E6B3E"
FONT="Arial"
DATA_ROWS=60  # rows reserved for tasks

thin=Side(style="thin", color="C9D2D8")
border=Border(left=thin,right=thin,top=thin,bottom=thin)

BUCKETS=["Pilot Path","Sales Pipeline","Business Infrastructure","Platform","Content"]
STATUSES=["Not Started","In Progress","Blocked","Ongoing","Done"]
PRIORITIES=["High","Medium","Low"]

HEADERS=["Bucket","Task","Status","Priority","Next Action","Target"]
WIDTHS=[22,52,14,10,52,16]

# ── Task snapshot (as of 2026-06-29) ──
# Cleared this reseed (Andrew confirmed 2026-06-29):
#   - Baird (MTPD) login — RESOLVED by the SMTP repoint.
#   - Confirm EGPD auth_uid linkage — CONFIRMED clean; completions saving.
#   - Repoint MTPD auth SMTP onto Google — DONE 6/26.
#   - Author EGPD per-module debrief legal summaries — DONE 6/29 (commit 0339eed).
#   - Shared-engine hardening pass + quiz change-answer — DONE 6/29, struck from board.
#   - EGPD badge artwork — REMOVED 6/29: badge (assets/egpd-badge.png) is live since
#     6/11 launch; the "waiting for correct artwork" blocker was unsubstantiated/stale.
TASKS=[
 ["Pilot Path",
  "Monitor + drive EGPD officer completion rates — the real lever on Chief Halteman signing",
  "Ongoing","High",
  "Weekly: check completions in Supabase; nudge Chief if engagement lags","Weekly"],
 ["Sales Pipeline",
  "Chief Halteman signs paid agreement (end of 3-month pilot)",
  "Not Started","High",
  "Outcome driven by completion rates above — convert pilot to paid","~2026-09-17"],
 ["Business Infrastructure",
  "Server-side content + quiz answer-key hardening — answer keys currently ship in public client JS",
  "Not Started","High",
  "HARD GATE: must close before Agency #3 / any PAYING agency goes live. Accepted risk for EGPD pilot only. Gate documented in Agency Onboarding/NEW-DEPARTMENT-ONBOARDING.md.","Pre-paid-launch"],
 ["Business Infrastructure",
  "Long-term email deliverability so future agencies' auto-emails inbox — egreenville.org whitelist of arbiterle.com sender + SPF/DKIM/DMARC review",
  "Not Started","Low",
  "EGPD pilot works around via manual Google sends (old 'Resend reputation' item dropped — Resend now used nowhere live)","Post-pilot"],
 ["Business Infrastructure",
  "Supabase key rotation (after welcome-email edge functions built)",
  "Not Started","Low",
  "Hygiene — sequence after deliverability work","Post-pilot"],
 ["Business Infrastructure",
  "Delete leftover MTPD 'TEST1' cleanup row from officers table",
  "Not Started","Low",
  "Pending Andrew's OK to delete the test row","—"],
 ["Platform",
  "Migrate inline MTPD debrief legal text out of shared app.js into MTPD module data (debriefHtml)",
  "Not Started","Low",
  "Tech debt — safe today via id-namespacing + smoke guard. Needs live-lock sign-off; removes the SMOKE-ALLOW-DEPT-NAMES allowlist.","Post-pilot"],
]

wb=Workbook()

# ── hidden Lists sheet (dropdown sources) ──
ls=wb.active; ls.title="Lists"
ls["A1"]="Buckets"; ls["B1"]="Statuses"; ls["C1"]="Priorities"
for i,v in enumerate(BUCKETS, start=2): ls.cell(row=i,column=1,value=v)
for i,v in enumerate(STATUSES, start=2): ls.cell(row=i,column=2,value=v)
for i,v in enumerate(PRIORITIES, start=2): ls.cell(row=i,column=3,value=v)
ls.sheet_state="hidden"

# ── Board ──
ws=wb.create_sheet("Board")
ncol=len(HEADERS)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
t=ws.cell(row=1,column=1,value="ARBITER LE  —  WORK BOARD")
t.font=Font(name=FONT,size=14,bold=True,color=GOLD)
t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[1].height=26

for c,(h,w) in enumerate(zip(HEADERS,WIDTHS), start=1):
    cell=ws.cell(row=2,column=c,value=h)
    cell.font=Font(name=FONT,size=10,bold=True,color=WHITE)
    cell.fill=PatternFill("solid",fgColor=STEEL)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=border
    ws.column_dimensions[get_column_letter(c)].width=w
ws.row_dimensions[2].height=24

WRAP_COLS={"Task","Next Action"}
r=3
for row in TASKS:
    for c,val in enumerate(row, start=1):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="top",
                                 wrap_text=(HEADERS[c-1] in WRAP_COLS))
        cell.border=border
    ws.row_dimensions[r].height=42
    r+=1
# Empty styled rows
for rr in range(r, 3+DATA_ROWS):
    for c in range(1,ncol+1):
        cell=ws.cell(row=rr,column=c)
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.border=border
        cell.alignment=Alignment(horizontal="left",vertical="top",
                                 wrap_text=(HEADERS[c-1] in WRAP_COLS))
    if rr%2==0:
        for c in range(1,ncol+1):
            ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=LIGHT)

last=2+DATA_ROWS

# Dropdowns: Bucket(A), Status(C), Priority(D)
dv_b=DataValidation(type="list",formula1="=Lists!$A$2:$A$%d"%(len(BUCKETS)+1),allow_blank=True)
dv_s=DataValidation(type="list",formula1="=Lists!$B$2:$B$%d"%(len(STATUSES)+1),allow_blank=True)
dv_p=DataValidation(type="list",formula1="=Lists!$C$2:$C$%d"%(len(PRIORITIES)+1),allow_blank=True)
for dv in (dv_b,dv_s,dv_p): ws.add_data_validation(dv)
dv_b.add("A3:A%d"%last); dv_s.add("C3:C%d"%last); dv_p.add("D3:D%d"%last)

# Conditional formatting — Status column (C) drives color
rng="C3:C%d"%last
ws.conditional_formatting.add(rng, FormulaRule(formula=['$C3="Blocked"'],
    fill=PatternFill("solid",fgColor=RED_FILL),
    font=Font(name=FONT,size=10,bold=True,color=RED_TXT)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$C3="In Progress"'],
    fill=PatternFill("solid",fgColor=GOLD_FILL)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$C3="Done"'],
    fill=PatternFill("solid",fgColor=GREEN_FILL),
    font=Font(name=FONT,size=10,strike=True,color=GREEN_TXT)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$C3="Ongoing"'],
    fill=PatternFill("solid",fgColor=BLUE_FILL)))
# Task text struck through when Done
ws.conditional_formatting.add("B3:B%d"%last, FormulaRule(formula=['$C3="Done"'],
    font=Font(name=FONT,size=10,strike=True,color="8A98A0")))
# High priority stands out
ws.conditional_formatting.add("D3:D%d"%last, FormulaRule(formula=['$D3="High"'],
    font=Font(name=FONT,size=10,bold=True,color=RED_TXT)))

ws.freeze_panes="A3"
ws.auto_filter.ref="A2:%s%d"%(get_column_letter(ncol),last)
ws.sheet_view.showGridLines=False

# ── Dashboard ──
db=wb.create_sheet("Dashboard")
db.sheet_view.showGridLines=False
for col,w in [("A",4),("B",30),("C",12),("D",4),("E",26),("F",12)]:
    db.column_dimensions[col].width=w
db.merge_cells("B1:F1")
h=db.cell(row=1,column=2,value="ARBITER LE  —  BOARD AT A GLANCE")
h.font=Font(name=FONT,size=14,bold=True,color=GOLD)
h.fill=PatternFill("solid",fgColor=NAVY)
h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
for c in range(2,7): db.cell(row=1,column=c).fill=PatternFill("solid",fgColor=NAVY)
db.row_dimensions[1].height=26

def section(cell,text):
    db[cell]=text
    db[cell].font=Font(name=FONT,size=11,bold=True,color=WHITE)
    db[cell].fill=PatternFill("solid",fgColor=STEEL)
    db[cell].alignment=Alignment(horizontal="left",indent=1)

def lab(rr,col,text,bold=False):
    cell=db.cell(row=rr,column=col,value=text)
    cell.font=Font(name=FONT,size=10,bold=bold,color=NAVY)
    return cell
def num(rr,col,formula,bold=True):
    cell=db.cell(row=rr,column=col,value=formula)
    cell.font=Font(name=FONT,size=10,bold=bold,color=NAVY)
    cell.alignment=Alignment(horizontal="center")
    return cell

# WIP CONTROL (the core rule, enforced)
section("B3","WIP Control"); section("C3","Now / Max")
wip_total_row=4
lab(wip_total_row,2,"In Progress — TOTAL",bold=True)
num(wip_total_row,3,'=COUNTIF(Board!$C$3:$C$%d,"In Progress")&" / 3"'%last)
rr=wip_total_row+1
for b in BUCKETS:
    lab(rr,2,"  In Progress — "+b)
    num(rr,3,'=COUNTIFS(Board!$A$3:$A$%d,"%s",Board!$C$3:$C$%d,"In Progress")&" / 1"'%(last,b,last))
    rr+=1
# Color-flag WIP breaches
db.conditional_formatting.add("C%d"%wip_total_row, FormulaRule(
    formula=['COUNTIF(Board!$C$3:$C$%d,"In Progress")>3'%last],
    fill=PatternFill("solid",fgColor=RED_FILL),font=Font(name=FONT,size=10,bold=True,color=RED_TXT)))
for i,b in enumerate(BUCKETS):
    cr=wip_total_row+1+i
    db.conditional_formatting.add("C%d"%cr, FormulaRule(
        formula=['COUNTIFS(Board!$A$3:$A$%d,"%s",Board!$C$3:$C$%d,"In Progress")>1'%(last,b,last)],
        fill=PatternFill("solid",fgColor=RED_FILL),font=Font(name=FONT,size=10,bold=True,color=RED_TXT)))

# By Status
section("E3","By Status"); section("F3","Count")
rr=4
for s in STATUSES:
    lab(rr,5,s); num(rr,6,'=COUNTIF(Board!$C$3:$C$%d,"%s")'%(last,s))
    rr+=1

# Watch list (below WIP block)
watch_start=wip_total_row+1+len(BUCKETS)+1
section("B%d"%watch_start,"Watch"); section("C%d"%watch_start,"Count")
watch=[("Blocked (needs unblocking)",'=COUNTIF(Board!$C$3:$C$%d,"Blocked")'%last),
       ("High priority, not done",'=COUNTIFS(Board!$D$3:$D$%d,"High",Board!$C$3:$C$%d,"<>Done")'%(last,last)),
       ("Active (not Done)",'=COUNTA(Board!$B$3:$B$%d)-COUNTIF(Board!$C$3:$C$%d,"Done")'%(last,last)),
       ("Missing a Next Action",'=SUMPRODUCT((Board!$B$3:$B$%d<>"")*(Board!$E$3:$E$%d=""))'%(last,last))]
rr=watch_start+1
for l,f in watch:
    lab(rr,2,l); num(rr,3,f); rr+=1

note=db.cell(row=rr+1,column=2,value="Counts recalc when the file opens. Review weekly (10 min).")
note.font=Font(name=FONT,size=8,italic=True,color=STEEL)
db.merge_cells("B%d:C%d"%(rr+1,rr+1))

# ── How to Use ──
hu=wb.create_sheet("How to Use")
hu.sheet_view.showGridLines=False
hu.column_dimensions["A"].width=2; hu.column_dimensions["B"].width=100
hb=hu.cell(row=2,column=2,value="HOW TO USE THIS BOARD")
hb.font=Font(name=FONT,size=14,bold=True,color=NAVY)
lines=[
 ("The rules (unchanged)",True),
 ("Max 3 In Progress total.  Max 1 In Progress per bucket.  Review weekly (10 min).  Blocked = write what's blocking it in the Next Action.",False),
 ("",False),
 ("The Dashboard enforces the rules",True),
 ("The WIP Control block shows In Progress now vs your max. The cell turns RED the moment you break a limit — total over 3, or any bucket over 1. If it's red, finish or park something before starting new work.",False),
 ("",False),
 ("Status drives the colors (automatic)",True),
 ("Blocked = red, In Progress = gold, Ongoing = blue, Done = green + struck through. You set the Status from the dropdown; the color follows. Bucket, Status, and Priority are all dropdowns.",False),
 ("",False),
 ("The columns",True),
 ("Bucket (which part of the business) · Task · Status · Priority · Next Action (the single next step) · Target (a date or a tag like 'Weekly' / 'Post-pilot'). Every active row should have a Next Action.",False),
 ("",False),
 ("Weekly habit (10 min)",True),
 ("1. Open the Dashboard. If WIP Control is red, fix it first. 2. Clear anything Blocked or High-priority. 3. Mark finished work Done (it strikes through — leave it or delete the row). 4. Make sure every active row still has a real Next Action.",False),
 ("",False),
 ("Confidentiality",True),
 ("Internal only. Caught by *.xlsx in .gitignore — never commits, never deploys. Never paste real officer training data here. Never record that EGPD's pilot was free where a prospect could see it.",False),
]
r=4
for text,is_head in lines:
    if text=="":
        r+=1; continue
    cell=hu.cell(row=r,column=2,value=text)
    if is_head:
        cell.font=Font(name=FONT,size=11,bold=True,color=GOLD)
    else:
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.alignment=Alignment(wrap_text=True,vertical="top")
        hu.row_dimensions[r].height=42
    r+=1

# Order tabs: Dashboard, Board, How to Use, Lists(hidden)
wb._sheets.sort(key=lambda s:{"Dashboard":0,"Board":1,"How to Use":2,"Lists":3}[s.title])
wb.active=wb.sheetnames.index("Dashboard")

out="2026-06-22-arbiter-le-tracker.xlsx"
wb.save(out)
print("saved",out)
