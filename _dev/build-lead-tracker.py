from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

NAVY="0A1828"; GOLD="C8902A"; STEEL="3C6478"
LIGHT="EFF2F4"; WHITE="FFFFFF"
RED_FILL="F8D7DA"; GOLD_FILL="FCEFC8"; GREEN_FILL="D7E8DC"
FONT="Arial"
DATA_ROWS=200  # rows reserved for prospects

thin=Side(style="thin", color="C9D2D8")
border=Border(left=thin,right=thin,top=thin,bottom=thin)

STAGES=["Identified","Researching","Contacted","In Conversation",
        "Demo Scheduled","Proposal Sent","Verbal Yes","Signed","Passed / Not Now"]
PRIORITIES=["High","Medium","Low"]

HEADERS=["Department","County / Area","Type","~Officers","Stage","Priority",
         "Contact Name","Title","Email","Phone","Source / Connection",
         "Last Touch","Next Action","Next Action Date","Notes"]
WIDTHS=[24,18,14,9,16,9,20,18,26,15,26,12,34,15,40]

wb=Workbook()

# ---------- hidden Lists sheet (dropdown sources) ----------
ls=wb.active; ls.title="Lists"
ls["A1"]="Stages"; ls["B1"]="Priorities"
for i,s in enumerate(STAGES, start=2): ls.cell(row=i,column=1,value=s)
for i,p in enumerate(PRIORITIES, start=2): ls.cell(row=i,column=2,value=p)
ls.sheet_state="hidden"

# ---------- Pipeline ----------
ws=wb.create_sheet("Pipeline")
ncol=len(HEADERS)
# Title row
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
t=ws.cell(row=1,column=1,value="ARBITER LE  —  LEAD PIPELINE")
t.font=Font(name=FONT,size=14,bold=True,color=GOLD)
t.fill=PatternFill("solid",fgColor=NAVY)
t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[1].height=26
# Header row
for c,(h,w) in enumerate(zip(HEADERS,WIDTHS), start=1):
    cell=ws.cell(row=2,column=c,value=h)
    cell.font=Font(name=FONT,size=10,bold=True,color=WHITE)
    cell.fill=PatternFill("solid",fgColor=STEEL)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=border
    ws.column_dimensions[get_column_letter(c)].width=w
ws.row_dimensions[2].height=30

# Seed rows: EGPD (real signed anchor) + one clearly-marked example
seed=[
 ["East Greenville PD","Montgomery (Borough)","Borough PD","","Signed","High",
  "Chief Halteman","Chief of Police","","","Pilot — signed agreement",
  "2026-06-17","Pilot live (free 3 mo) — convert to paid at pilot end","2026-09-17",
  "First paying agency. Pilot started 6/17. THE conversion to track: paid renewal at pilot end."],
 ["[Example — overwrite or delete this row]","Montgomery","Township PD","22","Contacted","Medium",
  "Jane Doe","Chief","jdoe@example.gov","(610) 555-0100","Cold — prospect list",
  "2026-06-20","Follow up call re: scheduling a demo","2026-06-27",
  "Replace with a real prospect. Set a Next Action + date for every live row."],
]
r=3
for row in seed:
    for c,val in enumerate(row, start=1):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.alignment=Alignment(horizontal="left",vertical="center",
                                 wrap_text=(HEADERS[c-1] in ("Next Action","Notes")))
        cell.border=border
        if HEADERS[c-1] in ("Last Touch","Next Action Date") and val:
            cell.number_format="yyyy-mm-dd"
    r+=1
# Empty styled rows
for rr in range(r, 3+DATA_ROWS):
    for c in range(1,ncol+1):
        cell=ws.cell(row=rr,column=c)
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.border=border
        cell.alignment=Alignment(horizontal="left",vertical="center",
                                 wrap_text=(HEADERS[c-1] in ("Next Action","Notes")))
        if HEADERS[c-1] in ("Last Touch","Next Action Date"):
            cell.number_format="yyyy-mm-dd"
    if rr%2==0:
        for c in range(1,ncol+1):
            ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=LIGHT)

last=2+DATA_ROWS
# Dropdowns
dv_stage=DataValidation(type="list",formula1="=Lists!$A$2:$A$%d"%(len(STAGES)+1),allow_blank=True)
dv_prio=DataValidation(type="list",formula1="=Lists!$B$2:$B$%d"%(len(PRIORITIES)+1),allow_blank=True)
ws.add_data_validation(dv_stage); ws.add_data_validation(dv_prio)
dv_stage.add("E3:E%d"%last); dv_prio.add("F3:F%d"%last)

# Conditional formatting on Next Action Date (col N=14): overdue red, due<=7d gold
ws.conditional_formatting.add("N3:N%d"%last,
    FormulaRule(formula=['AND($N3<>"",$N3<TODAY())'],
                fill=PatternFill("solid",fgColor=RED_FILL)))
ws.conditional_formatting.add("N3:N%d"%last,
    FormulaRule(formula=['AND($N3<>"",$N3>=TODAY(),$N3<=TODAY()+7)'],
                fill=PatternFill("solid",fgColor=GOLD_FILL)))
# Signed rows green on Stage
ws.conditional_formatting.add("E3:E%d"%last,
    FormulaRule(formula=['$E3="Signed"'],
                fill=PatternFill("solid",fgColor=GREEN_FILL)))

ws.freeze_panes="A3"
ws.auto_filter.ref="A2:%s%d"%(get_column_letter(ncol),last)
ws.sheet_view.showGridLines=False

# ---------- Dashboard ----------
db=wb.create_sheet("Dashboard")
db.sheet_view.showGridLines=False
for col,w in [("A",4),("B",24),("C",12),("D",4),("E",24),("F",12)]:
    db.column_dimensions[col].width=w
db.merge_cells("B1:F1")
h=db.cell(row=1,column=2,value="ARBITER LE  —  PIPELINE AT A GLANCE")
h.font=Font(name=FONT,size=14,bold=True,color=GOLD)
h.fill=PatternFill("solid",fgColor=NAVY)
h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
for c in range(2,7): db.cell(row=1,column=c).fill=PatternFill("solid",fgColor=NAVY)
db.row_dimensions[1].height=26

def section(cell,text):
    db[cell]=text
    db[cell].font=Font(name=FONT,size=11,bold=True,color=WHITE)
    db[cell].fill=PatternFill("solid",fgColor=STEEL)
    db[cell].alignment=Alignment(horizontal="left",indent=1)

# By Stage
section("B3","By Stage"); section("C3","Count")
for i,s in enumerate(STAGES):
    rr=4+i
    db.cell(row=rr,column=2,value=s).font=Font(name=FONT,size=10,color=NAVY)
    cc=db.cell(row=rr,column=3,value='=COUNTIF(Pipeline!$E$3:$E$%d,$B%d)'%(last,rr))
    cc.font=Font(name=FONT,size=10,color=NAVY)
    cc.alignment=Alignment(horizontal="center")
trow=4+len(STAGES)
db.cell(row=trow,column=2,value="Active leads (excl. Signed / Passed)").font=Font(name=FONT,size=10,bold=True,color=NAVY)
av=db.cell(row=trow,column=3,
   value='=COUNTA(Pipeline!$A$3:$A$%d)-COUNTIF(Pipeline!$E$3:$E$%d,"Signed")-COUNTIF(Pipeline!$E$3:$E$%d,"Passed / Not Now")'%(last,last,last))
av.font=Font(name=FONT,size=10,bold=True,color=NAVY); av.alignment=Alignment(horizontal="center")

# By Priority + Action flags
section("E3","Action Watch"); section("F3","Count")
flags=[("Overdue follow-ups",'=SUMPRODUCT((Pipeline!$N$3:$N$%d<>"")*(Pipeline!$N$3:$N$%d<TODAY()))'%(last,last)),
       ("Due in next 7 days",'=SUMPRODUCT((Pipeline!$N$3:$N$%d>=TODAY())*(Pipeline!$N$3:$N$%d<=TODAY()+7))'%(last,last)),
       ("Missing a Next Action",'=SUMPRODUCT((Pipeline!$A$3:$A$%d<>"")*(Pipeline!$M$3:$M$%d=""))'%(last,last)),
       ("High priority, open",'=COUNTIFS(Pipeline!$F$3:$F$%d,"High",Pipeline!$E$3:$E$%d,"<>Signed",Pipeline!$E$3:$E$%d,"<>Passed / Not Now")'%(last,last,last))]
for i,(lab,f) in enumerate(flags):
    rr=4+i
    db.cell(row=rr,column=5,value=lab).font=Font(name=FONT,size=10,color=NAVY)
    cc=db.cell(row=rr,column=6,value=f)
    cc.font=Font(name=FONT,size=10,bold=True,color=NAVY); cc.alignment=Alignment(horizontal="center")

note=db.cell(row=9,column=5,value="Refresh: data updates when the file opens / recalculates.")
note.font=Font(name=FONT,size=8,italic=True,color=STEEL)
db.merge_cells("E9:F9")

# ---------- How to Use ----------
hu=wb.create_sheet("How to Use")
hu.sheet_view.showGridLines=False
hu.column_dimensions["A"].width=2; hu.column_dimensions["B"].width=100
hu.merge_cells("B1:B1")
hb=hu.cell(row=2,column=2,value="HOW TO USE THIS TRACKER")
hb.font=Font(name=FONT,size=14,bold=True,color=NAVY)
lines=[
 ("","",False),
 ("The one rule","",True),
 ("Every live row has a Next Action and a Next Action Date. No exceptions. A lead with no next step is a lead you're losing. The Dashboard flags any row that breaks this rule.",None,False),
 ("","",False),
 ("The stages","",True),
 ("Identified → Researching → Contacted → In Conversation → Demo Scheduled → Proposal Sent → Verbal Yes → Signed.  (Passed / Not Now parks a dead lead without deleting the history.)",None,False),
 ("","",False),
 ("The colors (automatic)","",True),
 ("Next Action Date turns RED when it's overdue, GOLD when it's due within 7 days. Stage turns GREEN when Signed. You don't set these — they update themselves.",None,False),
 ("","",False),
 ("Weekly habit (15 min)","",True),
 ("1. Open the Dashboard. Clear every Overdue and Due-in-7-days item first.",None,False),
 ("2. Work the Active leads from highest stage down — closest to Signed gets your time.",None,False),
 ("3. After every touch: update Last Touch, write the next step, set its date. Advance the Stage if it moved.",None,False),
 ("","",False),
 ("Filling the pipeline","",True),
 ("Drop your prospect list into the Pipeline tab (one department per row). Stage and Priority are dropdowns. The EGPD row is the live anchor — its real Next Action is converting the pilot to paid at the end (tracked for 2026-09-17).",None,False),
 ("","",False),
 ("Confidentiality","",True),
 ("Internal only — this file is in .vercelignore and never deploys. Never paste real officer training data here. Never let a prospect see another agency's terms or that EGPD's pilot was free.",None,False),
]
r=4
for title,_,is_head in lines:
    if title=="":
        r+=1; continue
    cell=hu.cell(row=r,column=2,value=title)
    if is_head:
        cell.font=Font(name=FONT,size=11,bold=True,color=GOLD)
    else:
        cell.font=Font(name=FONT,size=10,color=NAVY)
        cell.alignment=Alignment(wrap_text=True,vertical="top")
        hu.row_dimensions[r].height=30
    r+=1

# Order tabs
wb.move_sheet("Dashboard", -(wb.sheetnames.index("Dashboard")))
wb._sheets.sort(key=lambda s:{"Dashboard":0,"Pipeline":1,"How to Use":2,"Lists":3}[s.title])
wb.active=wb.sheetnames.index("Dashboard")

out="_marketing/prospects/2026-06-22-arbiter-le-lead-tracker.xlsx"
wb.save(out)
print("saved",out)
